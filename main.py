import os
import logging
from flask import request, render_template, jsonify, send_file
from datetime import datetime
import openpyxl
import openai
from app import app, db
from models import MacroCategory, Macro

# ログ設定
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# OpenAI APIキーの設定
openai.api_key = os.getenv("OPENAI_API_KEY")

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/macros', methods=['GET'])
def get_macros():
    try:
        show_public = request.args.get('public', 'false').lower() == 'true'
        query = db.select(Macro)
        if show_public:
            query = query.filter_by(is_public=True)
        macros = db.session.execute(query.order_by(Macro.created_at.desc())).scalars().all()
        return jsonify([macro.to_dict() for macro in macros])
    except Exception as e:
        logger.error(f"Error in get_macros: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/generate-macro', methods=['POST'])
def generate_macro():
    try:
        data = request.get_json()
        template_id = data.get('template_id')
        use_ai = data.get('use_ai', True)
        description = data.get('description')
        category = data.get('category', MacroCategory.AI_GENERATED)

        logger.debug(f"Generating macro with parameters: template_id={template_id}, use_ai={use_ai}, category={category}")

        if template_id:
            template = db.session.execute(
                db.select(Macro).filter_by(id=template_id)
            ).scalar_one_or_none()

            if not template:
                return jsonify({"error": "Template not found"}), 404

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "マクロ"
            ws['A1'] = template.title
            ws['A2'] = template.description

            filename = f"macro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("temp", filename)
            os.makedirs("temp", exist_ok=True)
            wb.save(filepath)

            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        elif use_ai:
            if not description:
                return jsonify({"error": "Description is required for AI generation"}), 400

            logger.debug("Calling OpenAI API for macro generation")
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "あなたはExcelマクロの専門家です。ユーザーの要件に基づいて、実用的なVBAマクロを生成してください。"},
                    {"role": "user", "content": f"以下の要件に基づくExcelマクロを作成してください：\n{description}"}
                ]
            )
            generated_content = response.choices[0].message.content

            # マクロを保存
            macro = Macro(
                title=f"マクロ {datetime.now().strftime('%Y/%m/%d %H:%M')}",
                description=description,
                category=category
            )
            db.session.add(macro)
            db.session.commit()

            # バージョンを追加
            macro.add_version(generated_content)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "マクロ"
            ws['A1'] = macro.title
            ws['A2'] = generated_content

            filename = f"macro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("temp", filename)
            os.makedirs("temp", exist_ok=True)
            wb.save(filepath)

            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            return jsonify({"error": "Either template_id or description with use_ai must be provided"}), 400
    except Exception as e:
        logger.error(f"Error in generate_macro: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/macros/<int:macro_id>', methods=['GET'])
def get_macro(macro_id):
    try:
        macro = db.session.execute(db.select(Macro).filter_by(id=macro_id)).scalar_one_or_none()
        if macro:
            return jsonify(macro.to_dict())
        else:
            return jsonify({"error": "Macro not found"}), 404
    except Exception as e:
        logger.error(f"Error in get_macro: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/macros/<int:macro_id>/versions', methods=['GET'])
def get_macro_versions(macro_id):
    try:
        versions = db.session.execute(
            db.select(Macro.versions)
            .filter_by(id=macro_id)
            .order_by(Macro.versions.version_number.desc())
        ).scalars().all()
        return jsonify([version.to_dict() for version in versions])
    except Exception as e:
        logger.error(f"Error in get_macro_versions: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/macros/<int:macro_id>/share', methods=['POST'])
def share_macro(macro_id):
    try:
        macro = db.session.execute(db.select(Macro).filter_by(id=macro_id)).scalar_one_or_none()
        if not macro:
            return jsonify({"error": "Macro not found"}), 404

        data = request.get_json()
        is_public = data.get('is_public', True)
        macro.is_public = is_public
        db.session.commit()

        return jsonify({"message": "Sharing settings updated successfully"})
    except Exception as e:
        logger.error(f"Error in share_macro: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    logger.info("Starting Flask application...")
    app.run(host='0.0.0.0', port=5000, debug=True)